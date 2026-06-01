from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from database import get_db
import models, schemas, io
from auth_utils import require_admin
from email_utils import trigger_email
from datetime import datetime

router = APIRouter()

@router.post("/broadcast", response_model=schemas.MsgRes)
def broadcast(payload: schemas.BroadcastReq, bg: BackgroundTasks, current=Depends(require_admin), db: Session = Depends(get_db)):
    q = db.query(models.Student)
    if payload.course:
        q = q.filter(models.Student.course == payload.course)
    
    students = q.all()
    for s in students:
        # In-app notification
        db.add(models.Notification(student_id=s.id, message=f"📢 {payload.message}", target="student"))
        
        # Email notification
        subject = "📢 Important Announcement from Placement Cell"
        body = f"""
        <h2>Important Announcement</h2>
        <p>{payload.message}</p>
        <br>
        <p>Regards,<br>Placement Cell Team</p>
        """
        trigger_email(bg, subject, s.email, body)
        
    db.commit()
    return {"message": f"Broadcast sent to {len(students)} students."}

@router.get("/dashboard")
def dashboard(current=Depends(require_admin), db: Session = Depends(get_db)):
    placed  = db.query(models.Student).filter(models.Student.is_placed==True).count()
    total   = db.query(models.Student).count()
    return {
        "stats": {
            "total_students": total, "placed": placed, "unplaced": total-placed,
            "companies": db.query(models.Company).count(),
            "jobs": db.query(models.Job).count(),
            "applications": db.query(models.Application).count(),
            "shortlisted": db.query(models.Application).filter(models.Application.status=="Shortlisted").count(),
            "selected": db.query(models.Application).filter(models.Application.status=="Selected").count(),
        },
        "recent_placements": [
            {"student": a.student.name if a.student else "—", "job": a.job.title if a.job else "—",
             "company": a.job.company.name if (a.job and a.job.company) else "—", "salary": a.job.salary if a.job else "—"}
            for a in db.query(models.Application).options(joinedload(models.Application.student), joinedload(models.Application.job).joinedload(models.Job.company))
                .filter(models.Application.status=="Selected").limit(5).all()
        ],
        "recent_applications": [
            {"student": a.student.name if a.student else "—", "job": a.job.title if a.job else "—", "status": a.status, "date": str(a.applied_on)}
            for a in db.query(models.Application).options(joinedload(models.Application.student), joinedload(models.Application.job))
                .order_by(models.Application.created_at.desc()).limit(6).all()
        ]
    }

# ── INTERVIEW FORMS ────────────────────────────────────
@router.get("/interview-forms", response_model=List[schemas.IFormOut])
def list_forms(db: Session = Depends(get_db)):
    return db.query(models.InterviewForm).filter(models.InterviewForm.is_active == True).all()

@router.post("/interview-forms", response_model=schemas.IFormOut, status_code=201)
def create_form(payload: schemas.IFormCreate, bg: BackgroundTasks, current=Depends(require_admin), db: Session = Depends(get_db)):
    f = models.InterviewForm(**payload.model_dump())
    db.add(f); db.commit(); db.refresh(f)
    
    # Notify only relevant students (those who applied and are Shortlisted/Selected)
    relevant_apps = (
        db.query(models.Application)
        .options(joinedload(models.Application.student))
        .filter(models.Application.job_id == payload.job_id)
        .filter(models.Application.status.in_(["Applied", "Shortlisted", "Selected"]))
        .all()
    )
    
    j = db.query(models.Job).filter(models.Job.id == payload.job_id).first()
    co = db.query(models.Company).filter(models.Company.id == payload.company_id).first()
    
    for app in relevant_apps:
        s = app.student
        if not s: continue
        
        db.add(models.Notification(
            student_id=s.id, 
            message=f"Interview form available for {j.title if j else 'a job'} at {co.name if co else 'a company'} — check Interview Forms tab! 📝", 
            target="student"
        ))
        
        # Email Notification
        subject = f"Interview Form: {j.title if j else 'A Job'} at {co.name if co else 'a Company'}"
        body = f"<h2>Hello {s.name},</h2><p>A new interview form is available for you to fill out for <strong>{j.title}</strong> at <strong>{co.name}</strong>.</p><p>Please log in to your dashboard and check the 'Interview Forms' tab to access it.</p><br><p>Best,<br>Placement Cell</p>"
        trigger_email(bg, subject, s.email, body)
        
    db.commit()
    return f

@router.delete("/interview-forms/{fid}", response_model=schemas.MsgRes)
def del_form(fid: int, current=Depends(require_admin), db: Session = Depends(get_db)):
    f = db.query(models.InterviewForm).filter(models.InterviewForm.id == fid).first()
    if not f: raise HTTPException(status_code=404, detail="Not found.")
    db.delete(f); db.commit()
    return {"message": "Form deleted."}


# ── INTERVIEW FORM REPORT (Excel Download) ─────────────────────────
@router.get("/interview-forms/{form_id}/report")
def download_form_report(form_id: int, current=Depends(require_admin), db: Session = Depends(get_db)):
    form = db.query(models.InterviewForm).filter(models.InterviewForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Interview form not found.")

    job = db.query(models.Job).filter(models.Job.id == form.job_id).first()
    company = db.query(models.Company).filter(models.Company.id == form.company_id).first()

    all_applications = (
        db.query(models.Application)
        .options(joinedload(models.Application.student))
        .filter(models.Application.job_id == form.job_id)
        .all()
    )

    # Filtered applications for the detailed list (Applied/Shortlisted/Selected)
    applications = [a for a in all_applications if a.status in ["Applied", "Shortlisted", "Selected"]]

    # Summary counts
    stats = {
        "Applied": len(all_applications),
        "Shortlisted": len([a for a in all_applications if a.status == "Shortlisted"]),
        "Selected": len([a for a in all_applications if a.status == "Selected"]),
        "Rejected": len([a for a in all_applications if a.status == "Rejected"]),
    }

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants Report"

    # ── Title Row ──
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    job_name = job.title if job else "Job"
    co_name  = company.name if company else "Company"
    title_cell.value = f"Applicants Report — {job_name} @ {co_name}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1a1f3c", end_color="1a1f3c", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Summary Section ──
    ws.merge_cells("A2:C2")
    ws["A2"] = "Summary Statistics"
    ws["A2"].font = Font(bold=True, size=12)
    ws.row_dimensions[2].height = 20

    summary_headers = ["Status", "Count", "Percentage"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    summary_data = [
        ("Total Applied", stats["Applied"]),
        ("Shortlisted", stats["Shortlisted"]),
        ("Selected", stats["Selected"]),
        ("Rejected", stats["Rejected"]),
    ]

    for i, (label, count) in enumerate(summary_data, 4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=count)
        perc = (count / stats["Applied"] * 100) if stats["Applied"] > 0 else 0
        ws.cell(row=i, column=3, value=f"{perc:.1f}%")

    # ── Form URL info ──
    ws.merge_cells("E3:I3")
    ws["E3"] = f"Interview Form URL: {form.form_url}"
    ws["E3"].font = Font(italic=True, color="3b82f6", underline="single")
    
    ws.merge_cells("E4:I4")
    ws["E4"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["E4"].font = Font(size=9, color="666666")

    # ── Header Row ──
    headers = ["#", "Name", "Roll No", "Email", "Phone", "Course", "CGPA", "Applied On", "Status", "Resume Link"]
    accent_fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    
    header_row = 10
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="1a1f3c")
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 20

    # ── Data Rows ──
    status_colors = {"Selected": "10b981", "Shortlisted": "3b82f6", "Rejected": "ef4444", "Applied": "6b7280"}
    base_url = "http://localhost:8000" # Ideally from config
    for i, app in enumerate(applications, 1):
        s = app.student
        resume_url = f"{base_url}/uploads/resumes/{s.resume}" if (s and s.resume) else "—"
        
        row_data = [
            i,
            s.name    if s else "—",
            s.roll_no if s else "—",
            s.email   if s else "—",
            s.phone   if s else "—",
            s.course  if s else "—",
            s.cgpa    if s else "—",
            str(app.applied_on) if app.applied_on else "—",
            app.status,
            resume_url
        ]
        row_num = i + header_row
        row_fill = PatternFill(start_color="f8f9ff" if i % 2 == 0 else "FFFFFF", end_color="f8f9ff" if i % 2 == 0 else "FFFFFF", fill_type="solid")
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col in (1, 6, 7, 8, 9) else "left", vertical="center")
            cell.border = thin_border
            if col != 9:
                cell.fill = row_fill
                if col == 10 and val != "—":
                    cell.font = Font(color="3b82f6", underline="single")
            else:
                # Status column coloured
                sc = status_colors.get(app.status, "6b7280")
                cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # ── Column Widths ──
    col_widths = [5, 22, 14, 30, 14, 12, 8, 14, 14, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Stream response ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in job_name)
    filename = f"Applicants_{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ── APPLICANTS REPORT (Excel Download) ─────────────────────────
@router.get("/applicants/report")
def download_applicants_report(
    job_id: Optional[int] = None,
    status: Optional[str] = None,
    course: Optional[str] = None,
    current=Depends(require_admin),
    db: Session = Depends(get_db)
):
    q = db.query(models.Application).options(
        joinedload(models.Application.student),
        joinedload(models.Application.job).joinedload(models.Job.company)
    )
    
    if job_id: q = q.filter(models.Application.job_id == job_id)
    if status: q = q.filter(models.Application.status == status)
    if course: q = q.join(models.Student).filter(models.Student.course == course)
    
    all_apps = q.all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants List"

    # ── Title ──
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Applicants Data Report"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1a1f3c", end_color="1a1f3c", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Summary ──
    stats = {
        "Total": len(all_apps),
        "Shortlisted": len([a for a in all_apps if a.status == "Shortlisted"]),
        "Selected": len([a for a in all_apps if a.status == "Selected"]),
        "Rejected": len([a for a in all_apps if a.status == "Rejected"]),
        "Applied": len([a for a in all_apps if a.status == "Applied"]),
    }

    ws["A2"] = "Summary Stats"
    ws["A2"].font = Font(bold=True)
    ws.cell(row=3, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Count").font = Font(bold=True)
    
    summary_rows = [("Total", stats["Total"]), ("Applied", stats["Applied"]), ("Shortlisted", stats["Shortlisted"]), ("Selected", stats["Selected"]), ("Rejected", stats["Rejected"])]
    for i, (lbl, val) in enumerate(summary_rows, 4):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=val)

    ws.merge_cells("D2:F2")
    ws["D2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["D2"].font = Font(italic=True)

    # ── Headers ──
    header_row = 10
    headers = ["#", "Student Name", "Roll No", "Course", "CGPA", "Job Title", "Company", "Status", "Applied On", "Resume Link"]
    accent_fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Data ──
    status_colors = {"Selected": "10b981", "Shortlisted": "3b82f6", "Rejected": "ef4444", "Applied": "6b7280"}
    base_url = "http://localhost:8000"
    
    for i, app in enumerate(all_apps, 1):
        s = app.student
        j = app.job
        row_num = i + header_row
        resume_url = f"{base_url}/uploads/resumes/{s.resume}" if (s and s.resume) else "—"
        
        row_data = [
            i,
            s.name if s else "—",
            s.roll_no if s else "—",
            s.course if s else "—",
            s.cgpa if s else "—",
            j.title if j else "—",
            j.company.name if (j and j.company) else "—",
            app.status,
            str(app.applied_on) if app.applied_on else "—",
            resume_url
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if col == 8: # Status
                sc = status_colors.get(val, "6b7280")
                cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            if col == 10 and val != "—": # Resume Link
                cell.font = Font(color="3b82f6", underline="single")

    col_widths = [5, 25, 15, 12, 8, 25, 25, 15, 15, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Applicants_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
